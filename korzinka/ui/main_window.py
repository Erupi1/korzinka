from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QStackedWidget, QSizePolicy, QFileDialog, QMessageBox)
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import Qt
import os
from ui.product_table_widget import ProductTableWidget
from dialogs.add_product_dialog import AddProductDialog
from ui.expired_products_widget import ExpiredProductsWidget
from ui.removed_products_widget import RemovedProductsWidget
from dialogs.user_management_dialog import UserManagementDialog
from dialogs.change_password_dialog import ChangePasswordDialog
import openpyxl
from db import ProductDB

class MainWindow(QMainWindow):
    def __init__(self, user_info):
        super().__init__()
        self.user_info = user_info
        self.setWindowTitle('Korzinka — Мониторинг сроков годности')
        self.setMinimumSize(900, 600)
        self.init_ui()
        self.show_expiring_soon_popup()  # Показываем pop-up после инициализации UI

    def init_ui(self):
        main_widget = QWidget()
        main_layout = QHBoxLayout()
        # --- Sidebar ---
        sidebar = QVBoxLayout()
        sidebar.setSpacing(10)
        sidebar.setAlignment(Qt.AlignTop)
        # Кнопки sidebar (иконки можно заменить на png)
        self.btn_main = QPushButton('🏠 Главная')
        self.btn_expired = QPushButton('⏰ Просрочка')
        self.btn_expired.setObjectName('btn_expired')
        self.btn_add = QPushButton('➕ Добавить')
        self.btn_import = QPushButton('📥 Импорт')
        self.btn_export = QPushButton('📤 Экспорт')
        self.btn_remove = QPushButton('🗑️ Снять с полки')
        self.btn_history = QPushButton('📜 История')
        self.btn_password = QPushButton('🔑 Пароль')
        self.btn_users = QPushButton('👤 Пользователи')
        # Только для admin
        if self.user_info['role'] == 'admin':
            sidebar.addWidget(self.btn_users)
        # Добавляем кнопки
        for btn in [self.btn_main, self.btn_expired, self.btn_add, self.btn_import, self.btn_export, self.btn_remove, self.btn_history, self.btn_password, self.btn_users]:
            sidebar.addWidget(btn)
        sidebar.addStretch()
        # --- Основная рабочая область ---
        self.stack = QStackedWidget()
        # Страница "Главная" с таблицей товаров
        self.page_main = QWidget()
        main_layout_page = QVBoxLayout()
        self.product_table = ProductTableWidget()
        main_layout_page.addWidget(self.product_table)
        self.page_main.setLayout(main_layout_page)
        # Страница "Просрочка" с таблицей просроченных товаров
        self.page_expired = ExpiredProductsWidget(update_expired_count_callback=self.update_expired_count)
        # Страница "История" снятых товаров
        self.page_history = RemovedProductsWidget(is_admin=self.user_info['role'] == 'admin')
        self.stack.addWidget(self.page_main)
        self.stack.addWidget(self.page_expired)
        self.stack.addWidget(self.page_history)
        # --- Верхняя панель ---
        top_panel = QHBoxLayout()
        logo_path = os.path.join('resources', 'logo.png')
        logo = QLabel()
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path)
            logo.setPixmap(pixmap.scaled(40, 40, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            logo.setText('🛒')
            logo.setFixedSize(40, 40)
            logo.setAlignment(Qt.AlignCenter)
        title = QLabel('Корзинка — Мониторинг сроков годности')
        title.setProperty('role', 'title')
        title.setStyleSheet('font-size: 20px; font-weight: bold;')
        top_panel.addWidget(logo)
        top_panel.addSpacing(10)
        top_panel.addWidget(title)
        top_panel.addStretch()
        # --- Компоновка ---
        content_layout = QVBoxLayout()
        content_layout.addLayout(top_panel)
        content_layout.addWidget(self.stack)
        main_layout.addLayout(sidebar, 1)
        main_layout.addLayout(content_layout, 6)
        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)
        # --- Сигналы ---
        self.btn_main.clicked.connect(lambda: self.stack.setCurrentWidget(self.page_main))
        self.btn_expired.clicked.connect(lambda: self.stack.setCurrentWidget(self.page_expired))
        self.btn_add.clicked.connect(self.open_add_product_dialog)
        self.btn_history.clicked.connect(lambda: (self.page_history.load_data(), self.stack.setCurrentWidget(self.page_history)))
        self.btn_import.clicked.connect(self.import_products)
        self.btn_export.clicked.connect(self.export_products)
        self.btn_users.clicked.connect(self.open_user_management)
        self.btn_password.clicked.connect(self.open_change_password)
        self.update_expired_count()

    def update_expired_count(self):
        db = ProductDB()
        from PyQt5.QtCore import QDate
        today = QDate.currentDate().toString('yyyy-MM-dd')
        count = len(db.get_expired_products(today))
        if count > 0:
            self.btn_expired.setText(f'⏰ Просрочка ({count})')
        else:
            self.btn_expired.setText('⏰ Просрочка')

    def open_add_product_dialog(self):
        dialog = AddProductDialog(self)
        if dialog.exec_() == AddProductDialog.Accepted:
            self.product_table.load_data()
            self.page_expired.load_data()
            self.update_expired_count()

    def open_user_management(self):
        dialog = UserManagementDialog(self)
        dialog.exec_()

    def open_change_password(self):
        dialog = ChangePasswordDialog(self.user_info['id'], self)
        dialog.exec_()

    def import_products(self):
        path, _ = QFileDialog.getOpenFileName(self, 'Выберите Excel-файл', '', 'Excel Files (*.xlsx)')
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            db = ProductDB()
            count = 0
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                name, barcode, production_date, expiry_date, quantity, remind_date = row[:6]
                if not name or not production_date or not expiry_date or not quantity or not remind_date:
                    continue
                db.add_product(str(name), str(barcode) if barcode else None, str(production_date), str(expiry_date), int(quantity), str(remind_date))
                count += 1
            QMessageBox.information(self, 'Импорт завершён', f'Импортировано товаров: {count}')
            self.product_table.load_data()
            self.page_expired.load_data()
            self.update_expired_count()
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка импорта', str(e))

    def export_products(self):
        db = ProductDB()
        products = db.get_all_products()
        if not products:
            QMessageBox.information(self, 'Нет данных', 'Нет товаров для экспорта.')
            return
        path, _ = QFileDialog.getSaveFileName(self, 'Сохранить как', 'products.xlsx', 'Excel Files (*.xlsx)')
        if not path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['ID', 'Название', 'Штрихкод', 'Дата производства', 'Срок годности', 'Количество', 'Дата напоминания'])
        for p in products:
            ws.append(list(p[:7]))
        wb.save(path)
        QMessageBox.information(self, 'Успех', f'Экспортировано в {path}')

    def create_page(self, title, text, red=False):
        page = QWidget()
        layout = QVBoxLayout()
        label = QLabel(title)
        label.setStyleSheet(f'font-size: 18px; font-weight: bold; {"color: red;" if red else ""}')
        layout.addWidget(label)
        layout.addWidget(QLabel(text))
        layout.addStretch()
        page.setLayout(layout)
        return page

    def show_expiring_soon_popup(self):
        from PyQt5.QtCore import QDate
        db = ProductDB()
        today = QDate.currentDate().toString('yyyy-MM-dd')
        days = 7  # Можно вынести в настройки
        products = db.get_products_expiring_soon(today, days)
        if products:
            msg = '\n'.join([
                f"{p[1]} (штрихкод: {p[2] or '-'}), срок до: {p[4]}" for p in products
            ])
            QMessageBox.warning(self, f'Товары с истекающим сроком (до {days} дней)', msg) 